from io import BytesIO

from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook, load_workbook

from models import db, DropdownOption, Procurement
from routes.admin import is_admin

data_import_bp = Blueprint('data_import', __name__)

EXPORT_HEADERS = [
    '采购年份', '采购月份', '资产分类', '物品名称', '生产厂家', '物品型号',
    '预算数量', '预计单价', '总价', '需求部门', '需求人姓名', '需求人工号',
    '资产编码', '申请原因', '备注', '采购状态',
]

COLUMN_MAP = {
    '采购年份': 'year',
    '采购月份': 'month',
    '资产分类': 'asset_category',
    '资产最小分类': 'asset_category',
    '物品名称': 'item_name',
    '生产厂家': 'manufacturer',
    '物品型号': 'model',
    '预算数量': 'budget_qty',
    '预计单价': 'unit_price',
    '总价': 'total_price',
    '需求部门': 'department',
    '需求人姓名': 'requester_name',
    '需求人工号': 'requester_id',
    '资产编码': 'asset_code',
    '申请原因': 'reason',
    '备注': 'remark',
    '采购状态': 'status',
}

REQUIRED_FIELDS = [
    'year', 'month', 'asset_category', 'item_name', 'budget_qty',
    'unit_price', 'department', 'requester_name', 'requester_id', 'reason',
]


def build_option_sets():
    options = DropdownOption.query.all()
    grouped = {'asset_category': set(), 'department': set(), 'status': set()}
    for option in options:
        grouped.setdefault(option.category, set()).add(option.value)
    return grouped


def normalize_int(value):
    if value is None:
        return None
    text = str(value).replace('年', '').replace('月', '').strip()
    if not text:
        return None
    return int(float(text))


def normalize_float(value):
    if value is None:
        return None
    text = str(value).replace(',', '').strip()
    if not text:
        return None
    return float(text)


def is_blank_record(record):
    return all(value in (None, '') for value in record.values())


def validate_record(record, option_sets, row_idx):
    errors = []

    missing = [field for field in REQUIRED_FIELDS if record.get(field) in (None, '')]
    if missing:
        errors.append(f"缺少必填字段: {', '.join(missing)}")
        return errors

    if not 2022 <= record['year'] <= 2100:
        errors.append('采购年份不在合理范围内')
    if not 1 <= record['month'] <= 12:
        errors.append('采购月份必须在 1-12 之间')
    if record['budget_qty'] <= 0:
        errors.append('预算数量必须大于 0')
    if record['unit_price'] < 0:
        errors.append('预计单价不能小于 0')

    if option_sets['asset_category'] and record['asset_category'] not in option_sets['asset_category']:
        errors.append(f"资产分类不存在: {record['asset_category']}")
    if option_sets['department'] and record['department'] not in option_sets['department']:
        errors.append(f"需求部门不存在: {record['department']}")
    if record.get('status') and option_sets['status'] and record['status'] not in option_sets['status']:
        errors.append(f"采购状态不存在: {record['status']}")

    return errors


def procurement_to_row(item):
    return [
        item.year,
        item.month,
        item.asset_category,
        item.item_name,
        item.manufacturer or '',
        item.model or '',
        item.budget_qty,
        item.unit_price,
        item.total_price,
        item.department,
        item.requester_name,
        item.requester_id,
        item.asset_code or '',
        item.reason,
        item.remark or '',
        item.status,
    ]


@data_import_bp.route('/api/import/template', methods=['GET'])
def download_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = '采购导入模板'
    ws.append(EXPORT_HEADERS)
    ws.append([
        2026, 1, '办公设备', '示例物品', '示例厂家', '示例型号',
        1, 1000, 1000, '研发部', '张三', 'E20260001',
        'ASSET-001', '示例申请原因', '可选备注', '已申请',
    ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='采购导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@data_import_bp.route('/api/import/export-data', methods=['GET'])
def export_procurement_data():
    if not is_admin():
        return jsonify({'msg': '需要管理员权限'}), 403

    rows = Procurement.query.filter_by(is_deleted=0).order_by(
        Procurement.year.desc(), Procurement.month.desc(), Procurement.id.desc()
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '采购数据'
    ws.append(EXPORT_HEADERS)
    for row in rows:
        ws.append(procurement_to_row(row))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='采购数据导出.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@data_import_bp.route('/api/import/excel', methods=['POST'])
def import_excel():
    if not is_admin():
        return jsonify({'msg': '需要管理员权限'}), 403

    if 'file' not in request.files:
        return jsonify({'msg': '未找到文件'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'msg': '仅支持 .xlsx / .xls 格式'}), 400

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {idx: COLUMN_MAP[header] for idx, header in enumerate(headers) if header in COLUMN_MAP}
        if not col_map:
            return jsonify({'msg': '未匹配到有效列名，请先下载模板后再填写'}), 400

        option_sets = build_option_sets()
        imported = 0
        skipped_blank = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            record = {}
            for col_idx, field in col_map.items():
                val = row[col_idx].value if col_idx < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                    if not val:
                        val = None
                record[field] = val

            if is_blank_record(record):
                skipped_blank += 1
                continue

            row_errors = []

            try:
                record['year'] = normalize_int(record.get('year'))
            except Exception:
                row_errors.append('采购年份格式不正确')
                record['year'] = None

            try:
                record['month'] = normalize_int(record.get('month'))
            except Exception:
                row_errors.append('采购月份格式不正确')
                record['month'] = None

            try:
                record['budget_qty'] = int(normalize_float(record.get('budget_qty'))) if record.get('budget_qty') is not None else None
            except Exception:
                row_errors.append('预算数量格式不正确')
                record['budget_qty'] = None

            try:
                record['unit_price'] = normalize_float(record.get('unit_price'))
            except Exception:
                row_errors.append('预计单价格式不正确')
                record['unit_price'] = None

            total_price_raw = record.get('total_price')
            if total_price_raw not in (None, ''):
                try:
                    record['total_price'] = normalize_float(total_price_raw)
                except Exception:
                    row_errors.append('总价格式不正确')
                    record['total_price'] = None

            record['status'] = record.get('status') or '已申请'
            record['manufacturer'] = record.get('manufacturer') or ''
            record['model'] = record.get('model') or ''
            record['asset_code'] = record.get('asset_code') or ''
            record['remark'] = record.get('remark') or ''

            row_errors.extend(validate_record(record, option_sets, row_idx))

            if row_errors:
                errors.append(f"第 {row_idx} 行：{'；'.join(row_errors)}")
                continue

            record['total_price'] = round(record['budget_qty'] * record['unit_price'], 2)
            db.session.add(Procurement(**record))
            imported += 1

        db.session.commit()
        wb.close()

        return jsonify({
            'imported': imported,
            'skipped_blank': skipped_blank,
            'errors': errors[:30],
            'total_errors': len(errors),
            'msg': '导入完成',
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'msg': f'导入失败: {str(e)}'}), 500
